import subprocess
import sys
import os
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Global verbose flag
VERBOSE = False

def debug(msg):
    if VERBOSE:
        print(f"[DEBUG] {msg}")

# Function to check if required modules are installed
def check_dependencies():
    try:
        import openpyxl
    except ImportError:
        print("Error: 'openpyxl' is not installed. Please install it using: pip install openpyxl")
        sys.exit(1)

    # Built-ins should always be present, but we check for consistency
    for module in ('subprocess', 're', 'os'):
        try:
            __import__(module)
        except ImportError:
            print(f"Error: Required module {module} is not installed.")
            sys.exit(1)

# Ensure dependencies are available
check_dependencies()

def extract_cipher_suites(pcap_file, intermediate_file="cipher_suites_frm_pcap.txt"):
    """
    Runs tshark on the given pcap to extract raw TLS handshake ciphersuite data.
    Appends results to intermediate_file.
    """
    debug(f"Starting extract_cipher_suites for {pcap_file}")
    with open(intermediate_file, "a") as f:
        f.write(f"\n===== {os.path.basename(pcap_file)} =====\n")

    tshark_cmd = [
        "tshark",
        "-r", pcap_file,
        "-Y", "ssl.handshake.ciphersuites",
        "-Vx"
    ]
    debug(f"Running command: {' '.join(tshark_cmd)}")
    with open(intermediate_file, "a") as f:
        subprocess.run(tshark_cmd, stdout=f, stderr=subprocess.DEVNULL)
    debug(f"Completed tshark extraction to {intermediate_file}")


def extract_cipher_blocks(input_file="cipher_suites_frm_pcap.txt", output_file="extracted_cipher_suites.txt"):
    """
    Extracts blocks of cipher suites from tshark output,
    filters out GREASE entries, deduplicates, and writes to output_file.
    """
    debug(f"Starting extract_cipher_blocks from {input_file}")
    text = open(input_file).read()

    # Capture raw blocks before extension server_name
    pattern = re.compile(
        r"(Cipher Suites Length: .*?)(?=Type: server_name \(0\))",
        re.DOTALL | re.MULTILINE
    )
    matches = pattern.findall(text)
    debug(f"Found {len(matches)} raw blocks via regex")

    seen = set()
    unique_blocks = []
    for i, block in enumerate(matches, 1):
        # Remove any lines containing GREASE
        filtered = [ln for ln in block.splitlines() if 'Reserved (GREASE)' not in ln]
        block_clean = "\n".join(filtered).strip()
        if block_clean not in seen:
            seen.add(block_clean)
            unique_blocks.append(block_clean)
            debug(f"Adding unique cleaned block #{len(unique_blocks)} from raw block #{i}")
        else:
            debug(f"Skipping duplicate cleaned block from raw block #{i}")

    with open(output_file, "w") as f:
        for blk in unique_blocks:
            f.write(blk + "\n\n" + "="*40 + "\n\n")
    debug(f"Wrote {len(unique_blocks)} unique blocks to {output_file}")


def extract_domain_name(block_text):
    """
    Extracts the domain name from a cipher suite block.
    """
    match = re.search(r'Extension: server_name \(len=\d+\) name=([^\s]+)', block_text)
    return match.group(1) if match else ""


def parse_blocks(file_path):
    """
    Parses cleaned cipher suite blocks, removes GREASE entries,
    and collects server names.

    Returns:
      - blocks: list of cleaned block text
      - server_names: set of unique server names
    """
    debug(f"Starting parse_blocks on {file_path}")
    content = open(file_path).read()
    raw_blocks = content.strip().split("========================================")
    debug(f"Split into {len(raw_blocks)} raw blocks")

    server_names = set()
    name_pattern = re.compile(r'Extension: server_name \(len=\d+\) name=([^\s]+)')
    cleaned_blocks = []

    for idx, raw in enumerate(raw_blocks, 1):
        debug(f"Processing block {idx}")
        filtered = [ln for ln in raw.splitlines() if 'Reserved (GREASE)' not in ln]
        debug(f"Block {idx}: {len(raw.splitlines())} lines, {len(filtered)} after GREASE filter")

        cipher_suites = [line.split('Cipher Suite:')[1].strip() for line in filtered if line.strip().startswith('Cipher Suite:')]
        extension_lines = [line.strip() for line in filtered if line.strip().startswith('Extension: server_name')]
        if extension_lines and cipher_suites:
            server_name = name_pattern.search(extension_lines[0]).group(1)
            server_names.add(server_name)
            debug(f"Block {idx}: extracted server name '{server_name}' and {len(cipher_suites)} suites")

            block_text = "\n".join(filtered).strip()
            cleaned_blocks.append(block_text)
            debug(f"Block {idx}: added cleaned block")
        else:
            debug(f"Block {idx}: skipped (missing suites or extension)")

    debug(f"parse_blocks done: {len(cleaned_blocks)} cleaned blocks, {len(server_names)} unique server names")
    return cleaned_blocks, server_names


def export_to_excel(blocks, server_names, excel_file="output.xlsx"):
    """
    Exports cipher suite blocks and server names to an Excel workbook,
    sets column B width and autofits row height.
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Cipher Suite Blocks"

    # Set column widths
    ws1.column_dimensions['B'].width = 64  # Approx. 450 pixels
    ws1.column_dimensions['C'].width = 30  # Adjust as needed for domain names

    # Write headers
    ws1.append(["Block Index", "Cipher Suite Block", "Domain Name"])

    # Write blocks and corresponding domain names
    for i, blk in enumerate(blocks, 1):
        # Extract domain name from the block
        domain_name = extract_domain_name(blk)
        ws1.append([i, blk, domain_name])

    # Server Names sheet
    ws2 = wb.create_sheet(title="Server Names")
    ws2.append(["Index", "Server Name"])
    for i, name in enumerate(sorted(server_names), 1):
        ws2.append([i, name])

    # Style headers
    for ws in (ws1, ws2):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb.save(excel_file)
    print(f"Excel output written to '{excel_file}'")



def cleanup_files(files):
    for f in files:
        if os.path.exists(f):
            os.remove(f)
            print(f"Deleted file: {f}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Extract TLS cipher suites from a PCAP and export to Excel",
        usage="%(prog)s [-v] <pcap_file>"
    )
    parser.add_argument(
        "pcap_file", 
        help="Path to the input PCAP file"
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable verbose debug output"
    )
    args = parser.parse_args()

    # Enable verbose if requested
    VERBOSE = args.verbose

    pcap = args.pcap_file
    if not os.path.isfile(pcap):
        print(f"File not found: {pcap}")
        sys.exit(1)

    out_dir = os.path.splitext(os.path.basename(pcap))[0]
    os.makedirs(out_dir, exist_ok=True)

    intermediate_txt = "cipher_suites_frm_pcap.txt"
    extracted_txt = "extracted_cipher_suites.txt"

    # 1. Extract raw data
    extract_cipher_suites(pcap)
    # 2. Clean, filter GREASE and dedupe blocks
    extract_cipher_blocks()
    # 3. Parse blocks and get server names
    blocks, names = parse_blocks(extracted_txt)
    # 4. Export to Excel in output directory
    excel_path = os.path.join(out_dir, "output.xlsx")
    export_to_excel(blocks, names, excel_path)
    # 5. Cleanup
    cleanup_files([intermediate_txt, extracted_txt])

    print(f"Process completed. Final output in '{excel_path}'")
