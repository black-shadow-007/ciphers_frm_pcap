import subprocess
import sys
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Function to check if required modules are installed
def check_dependencies():
    try:
        import openpyxl
    except ImportError:
        print("Error: 'openpyxl' is not installed. Please install it using: pip install openpyxl")
        sys.exit(1)

    # Additional checks for other required dependencies
    try:
        import subprocess
        import re
        import os
    except ImportError as e:
        print(f"Error: Required module {e.name} is not installed.")
        sys.exit(1)

# Check dependencies before proceeding
check_dependencies()

def extract_cipher_suites(pcap_file, intermediate_file="cipher_suites_frm_pcap.txt"):
    with open(intermediate_file, "a") as f:
        f.write(f"\n===== {os.path.basename(pcap_file)} =====\n")

    tshark_cmd = [
        "tshark",
        "-r", pcap_file,
        "-Y", "ssl.handshake.ciphersuites",
        "-Vx"
    ]

    with open(intermediate_file, "a") as f:
        subprocess.run(tshark_cmd, stdout=f, stderr=subprocess.DEVNULL)

def extract_cipher_blocks(input_file="cipher_suites_frm_pcap.txt", output_file="extracted_cipher_suites.txt"):
    with open(input_file, "r") as f:
        text = f.read()

    pattern = re.compile(
        r"(Cipher Suites Length: .*?)(?=Type: server_name \(0\))",
        re.DOTALL | re.MULTILINE
    )

    blocks = pattern.findall(text)

    seen = set()
    unique_blocks = []
    for block in blocks:
        block_clean = block.strip()
        if block_clean not in seen:
            seen.add(block_clean)
            unique_blocks.append(block_clean)

    with open(output_file, "w") as f:
        for block in unique_blocks:
            f.write(block + "\n\n" + "="*40 + "\n\n")

    print(f"Extracted {len(unique_blocks)} unique cipher suite block(s) to {output_file}")

def parse_blocks(file_path):
    with open(file_path, "r") as file:
        content = file.read()

    blocks = content.strip().split("========================================")
    parsed_blocks = []
    server_names = set()

    server_name_pattern = re.compile(r'Extension: server_name \(len=\d+\) name=([^\s]+)')

    for block in blocks:
        lines = block.strip().splitlines()
        cipher_suites = []
        extension_line = ""
        server_name = ""

        for line in lines:
            line = line.strip()
            if line.startswith("Cipher Suite:"):
                cipher = line.split("Cipher Suite:")[1].strip()
                cipher_suites.append(cipher)
            elif "Extension: server_name" in line:
                extension_line = line.strip()
                match = server_name_pattern.search(line)
                if match:
                    server_name = match.group(1)
                    server_names.add(server_name)

        if cipher_suites and extension_line:
            parsed_blocks.append((frozenset(cipher_suites), extension_line, block.strip()))

    return parsed_blocks, server_names

def deduplicate_blocks(parsed_blocks):
    seen = set()
    unique_blocks = []

    for cipher_suites, extension, block in parsed_blocks:
        key = (cipher_suites, extension)
        if key not in seen:
            seen.add(key)
            unique_blocks.append(block)

    return unique_blocks

def export_to_excel(blocks, server_names, excel_file="output.xlsx"):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Cipher Suite Blocks"

    # Cipher Suite Blocks Sheet
    ws1.append(["Block Index", "Cipher Suite Block"])
    for i, block in enumerate(blocks, 1):
        ws1.append([i, block])
        ws1.row_dimensions[i + 1].height = 60  # Increase row height for better readability

    # Server Names Sheet
    ws2 = wb.create_sheet(title="Server Names")
    ws2.append(["Index", "Server Name"])
    for i, name in enumerate(sorted(server_names), 1):
        ws2.append([i, name])

    # Style headers
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb.save(excel_file)
    print(f"Excel output written to '{excel_file}'")

def cleanup_files(intermediate_files):
    for file in intermediate_files:
        if os.path.exists(file):
            os.remove(file)
            print(f"Deleted file: {file}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python extract_and_parse_tls.py <capture.pcap>")
        sys.exit(1)

    pcap_file = sys.argv[1]
    if not os.path.isfile(pcap_file):
        print(f"File not found: {pcap_file}")
        sys.exit(1)

    # Create output directory with the same name as pcap file (without extension)
    output_dir = os.path.splitext(os.path.basename(pcap_file))[0]
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    intermediate_files = ["cipher_suites_frm_pcap.txt", "extracted_cipher_suites.txt"]

    # Step 1: Extract raw cipher suite data from the pcap
    extract_cipher_suites(pcap_file)

    # Step 2: Extract individual cipher suite blocks
    extract_cipher_blocks()

    # Step 3: Parse and deduplicate
    parsed_blocks, server_names = parse_blocks("extracted_cipher_suites.txt")
    unique_blocks = deduplicate_blocks(parsed_blocks)

    # Step 4: Export to Excel in the new directory
    excel_file = os.path.join(output_dir, "output.xlsx")
    export_to_excel(unique_blocks, server_names, excel_file)

    # Step 5: Cleanup intermediate files
    cleanup_files(intermediate_files)

    print(f"Process completed. Final output saved in '{output_dir}/output.xlsx'")
